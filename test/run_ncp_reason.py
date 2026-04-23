"""
基于 test/ncp_test_0423.csv 中的"问题"列作为测试集，
并发请求 /api/reason 接口，并将推理结果实时（边完成边写入）保存到 CSV。

请求体只包含 question 与 policyId 两个字段，其它参数走服务端默认值。

用法示例（PowerShell）：
    python test/run_ncp_reason.py `
        --input test/ncp_test_0423.csv `
        --output test/ncp_test_0423_result.csv `
        --url http://mlp.paas.dc.servyou-it.com/kh-server/api/reason `
        --policy-id KH1493204307733168128_20260423163645 `
        --concurrency 5 `
        --timeout 600
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import requests


DEFAULT_URL = "http://mlp.paas.dc.servyou-it.com/kh-server/api/reason"
DEFAULT_POLICY_ID = "KH1493204307733168128_20260423163645"
DEFAULT_INPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ncp_test_0423.csv")
DEFAULT_OUTPUT = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    f"ncp_test_0423_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
)

OUTPUT_FIELDS = [
    "idx",
    "question",
    "原因",
    "缺陷",
    "http_status",
    "biz_status",
    "biz_message",
    "elapsed_sec",
    "answer",
    "think",
    "khObj",
    "skillsResult",
    "error",
    "finished_at",
]

EXTRA_INPUT_COLUMNS = ["原因", "缺陷"]


def load_questions(path: str) -> list[dict]:
    """读取测试集"问题"列（必填），并附带 EXTRA_INPUT_COLUMNS 中可选的额外列。

    返回 [{"question": str, "原因": str, "缺陷": str, ...}]，缺失列自动以空串填充。
    支持 .csv 与 .xlsx。
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        try:
            import openpyxl  # type: ignore
        except ImportError as e:
            raise RuntimeError("读取 xlsx 需要 openpyxl，请先 pip install openpyxl") from e
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        header = [
            str(c.value).strip() if c.value is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        if "问题" not in header:
            raise ValueError(f"输入 xlsx 缺少'问题'列，实际表头={header}")
        q_idx = header.index("问题")
        extra_idx = {col: (header.index(col) if col in header else -1) for col in EXTRA_INPUT_COLUMNS}
        items: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if q_idx >= len(row):
                continue
            v = row[q_idx]
            if v is None:
                continue
            q = str(v).strip()
            if not q:
                continue
            item = {"question": q}
            for col, ci in extra_idx.items():
                val = row[ci] if 0 <= ci < len(row) else None
                item[col] = "" if val is None else str(val).strip()
            items.append(item)
        return items

    items = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if "问题" not in (reader.fieldnames or []):
            raise ValueError(f"输入 CSV 缺少'问题'列，实际表头={reader.fieldnames}")
        for row in reader:
            q = (row.get("问题") or "").strip()
            if not q:
                continue
            item = {"question": q}
            for col in EXTRA_INPUT_COLUMNS:
                val = row.get(col)
                item[col] = "" if val is None else str(val).strip()
            items.append(item)
    return items


def call_reason(
    url: str,
    policy_id: str,
    question: str,
    timeout: float,
    extra_payload: dict | None = None,
) -> dict:
    """单次调用 /api/reason，返回标准化结果 dict（包含失败信息）。

    extra_payload：在基础入参（question/policyId）之上附加的字段，
    例如 {"answerSystemPrompt": "..."}；不会覆盖基础字段。
    """
    payload = {"question": question, "policyId": policy_id}
    if extra_payload:
        for k, v in extra_payload.items():
            if k not in payload:
                payload[k] = v
    started = time.time()
    result: dict = {
        "http_status": None,
        "biz_status": None,
        "biz_message": "",
        "elapsed_sec": 0.0,
        "answer": "",
        "think": "",
        "khObj": "",
        "skillsResult": "",
        "error": "",
    }
    try:
        resp = requests.post(
            url,
            json=payload,
            timeout=timeout,
            headers={"Content-Type": "application/json"},
        )
        result["http_status"] = resp.status_code
        result["elapsed_sec"] = round(time.time() - started, 3)
        try:
            body = resp.json()
        except Exception:
            result["error"] = f"响应非 JSON: {resp.text[:500]}"
            return result

        result["biz_status"] = body.get("status_code")
        result["biz_message"] = body.get("message", "")
        data = body.get("data") or {}
        if isinstance(data, dict):
            result["answer"] = data.get("answer", "") or ""
            result["think"] = data.get("think", "") or ""
            result["khObj"] = data.get("khObj", "") or ""
            skills = data.get("skillsResult") or {}
            if skills:
                try:
                    result["skillsResult"] = json.dumps(skills, ensure_ascii=False)
                except Exception:
                    result["skillsResult"] = str(skills)
    except requests.Timeout:
        result["elapsed_sec"] = round(time.time() - started, 3)
        result["error"] = f"请求超时(>{timeout}s)"
    except Exception as e:
        result["elapsed_sec"] = round(time.time() - started, 3)
        result["error"] = f"{type(e).__name__}: {e}"
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="并发跑 ncp 测试集并实时写入 CSV")
    parser.add_argument("--input", default=DEFAULT_INPUT, help="输入 CSV 路径")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="输出 CSV 路径")
    parser.add_argument("--url", default=DEFAULT_URL, help="/api/reason 接口地址")
    parser.add_argument("--policy-id", default=DEFAULT_POLICY_ID, help="policyId")
    parser.add_argument("--concurrency", type=int, default=5, help="并发线程数")
    parser.add_argument("--timeout", type=float, default=600.0, help="单次请求超时(s)")
    parser.add_argument("--limit", type=int, default=0, help="只跑前 N 条（0=全部）")
    parser.add_argument(
        "--answer-system-prompt",
        default="",
        help="answerSystemPrompt 字段内容（命令行直传，与 --answer-system-prompt-file 二选一）",
    )
    parser.add_argument(
        "--answer-system-prompt-file",
        default="",
        help="从文件读取 answerSystemPrompt 内容（utf-8）",
    )
    parser.add_argument("--vendor", default="", help="LLM vendor，如 aliyun / qwen3.5-122b-a10b；为空走服务端默认")
    parser.add_argument("--model", default="", help="LLM model 名，如 deepseek-v3.2；为空走服务端默认")
    parser.add_argument(
        "--extra-json",
        default="",
        help='额外透传到请求体的字段，JSON 字符串。例如 \'{"reduceMaxPartDepth":4,"relationMaxNodes":999}\'',
    )
    args = parser.parse_args()

    extra_payload: dict = {}
    if args.answer_system_prompt_file:
        with open(args.answer_system_prompt_file, "r", encoding="utf-8") as fp:
            extra_payload["answerSystemPrompt"] = fp.read()
    elif args.answer_system_prompt:
        extra_payload["answerSystemPrompt"] = args.answer_system_prompt
    if args.vendor:
        extra_payload["vendor"] = args.vendor
    if args.model:
        extra_payload["model"] = args.model
    if args.extra_json:
        try:
            extras = json.loads(args.extra_json)
        except Exception as e:
            print(f"[ERROR] 解析 --extra-json 失败: {e}")
            return 1
        if not isinstance(extras, dict):
            print("[ERROR] --extra-json 必须是 JSON 对象")
            return 1
        for k, v in extras.items():
            extra_payload[k] = v

    questions = load_questions(args.input)
    if args.limit > 0:
        questions = questions[: args.limit]
    total = len(questions)
    if total == 0:
        print("[ERROR] 输入 CSV 中没有有效问题")
        return 1

    out_ext = os.path.splitext(args.output)[1].lower()
    final_output = args.output
    if out_ext == ".xlsx":
        csv_output = os.path.splitext(args.output)[0] + ".csv"
    else:
        csv_output = args.output

    print(f"[INFO] 输入文件: {args.input}")
    print(f"[INFO] 输出文件: {final_output}" + (f"（过程实时写 {csv_output}）" if out_ext == ".xlsx" else ""))
    print(f"[INFO] 接口地址: {args.url}")
    print(f"[INFO] policyId : {args.policy_id}")
    print(f"[INFO] 总问题数 : {total}, 并发: {args.concurrency}, 超时: {args.timeout}s")
    if extra_payload:
        keys = ", ".join(f"{k}({len(str(v))}字符)" for k, v in extra_payload.items())
        print(f"[INFO] 额外入参 : {keys}")

    os.makedirs(os.path.dirname(os.path.abspath(final_output)) or ".", exist_ok=True)

    write_lock = threading.Lock()
    counter = {"done": 0, "ok": 0, "fail": 0}
    overall_start = time.time()

    with open(csv_output, "w", encoding="utf-8-sig", newline="") as fout:
        writer = csv.DictWriter(fout, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        fout.flush()

        def worker(idx: int, item: dict) -> dict:
            question = item["question"]
            res = call_reason(
                args.url,
                args.policy_id,
                question,
                args.timeout,
                extra_payload=extra_payload or None,
            )
            row = {
                "idx": idx,
                "question": question,
                "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                **{col: item.get(col, "") for col in EXTRA_INPUT_COLUMNS},
                **res,
            }
            with write_lock:
                writer.writerow(row)
                fout.flush()
                try:
                    os.fsync(fout.fileno())
                except OSError:
                    pass
                counter["done"] += 1
                ok = res["biz_status"] == 200 and not res["error"]
                if ok:
                    counter["ok"] += 1
                else:
                    counter["fail"] += 1
                preview = (res["answer"] or res["error"] or "").replace("\n", " ")[:60]
                print(
                    f"[{counter['done']}/{total}] "
                    f"idx={idx} ok={ok} biz={res['biz_status']} "
                    f"http={res['http_status']} {res['elapsed_sec']}s | {preview}",
                    flush=True,
                )
            return row

        with ThreadPoolExecutor(max_workers=args.concurrency) as pool:
            futures = [
                pool.submit(worker, i + 1, item) for i, item in enumerate(questions)
            ]
            for fut in as_completed(futures):
                try:
                    fut.result()
                except Exception as e:
                    with write_lock:
                        counter["done"] += 1
                        counter["fail"] += 1
                    print(f"[WARN] worker 异常: {type(e).__name__}: {e}", flush=True)

    elapsed = time.time() - overall_start
    print(
        f"\n[DONE] 总数={total} 成功={counter['ok']} 失败={counter['fail']} "
        f"总耗时={elapsed:.1f}s 平均={elapsed/total:.2f}s/题"
    )
    print(f"[DONE] CSV 已写入: {csv_output}")

    if out_ext == ".xlsx":
        try:
            _csv_to_xlsx(csv_output, final_output)
            print(f"[DONE] XLSX 已生成: {final_output}")
        except Exception as e:
            print(f"[WARN] CSV → XLSX 转换失败: {type(e).__name__}: {e}（CSV 仍可用）")
    return 0


def _csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """把已写好的 CSV 转成 XLSX。每个单元格按字符串写入，避免数字/日期被自动解析。"""
    try:
        import openpyxl  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError as e:
        raise RuntimeError("生成 xlsx 需要 openpyxl，请先 pip install openpyxl") from e

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="result")
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i == 0:
                ws.append(row)
            else:
                ws.append([("" if c is None else str(c)) for c in row])
    wb.save(xlsx_path)


if __name__ == "__main__":
    sys.exit(main())
