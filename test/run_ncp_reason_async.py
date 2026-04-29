"""
使用 /api/reason/submit + /api/reason/result/{taskId} 的**异步模式**跑 ncp 测试集。

相比 run_ncp_reason.py 的同步 /api/reason 调用：
- submit 阶段只做"入队 + 立刻返回 taskId"，连接短平快，不会被中间代理长连接超时掐断。
- submit 完成后 taskId 立刻落盘到结果 CSV（即使脚本中途被杀也不会丢失 taskId，
  后续可用 --poll-only 继续查结果）。
- 随后并发轮询 /api/reason/result/{taskId}，拿到 done/failed 即把完整结果回写。

用法示例（PowerShell）：
    python test/run_ncp_reason_async.py `
        --input test/ncp_test_0423.csv `
        --output test/ncp_test_0423_result_async.csv `
        --base-url http://mlp.paas.dc.servyou-it.com/kh-server/api/reason `
        --policy-id KH1498620964263747584_20260428113440 `
        --submit-concurrency 10 `
        --poll-concurrency 20 `
        --poll-interval 2 `
        --poll-timeout 3600

分阶段：
    # 只提交
    python test/run_ncp_reason_async.py --submit-only --output <csv>
    # 后续基于已经记录的 taskId 继续查结果
    python test/run_ncp_reason_async.py --poll-only   --output <csv>
    # 对结果表中 status≠done 的行重新 submit + poll，写回原 CSV
    python test/run_ncp_reason_async.py --retry-failed --output <csv>
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import requests


DEFAULT_BASE_URL = "http://mlp.paas.dc.servyou-it.com/kh-server/api/reason"
DEFAULT_POLICY_ID = "KH1498620964263747584_20260428113440"
DEFAULT_INPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ncp_test_0423.csv")
DEFAULT_OUTPUT = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    f"ncp_test_0423_result_async_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
)

# 完整输出字段：先写 submit 相关，再写 poll 拿到的最终结果
OUTPUT_FIELDS = [
    "idx",
    "question",
    "taskId",
    "submit_http_status",
    "submit_biz_status",
    "submit_error",
    "submitted_at",
    # 以下字段由 poll 阶段填充
    "status",          # pending / running / done / failed / timeout / not_found
    "enqueue_time",
    "start_time",
    "end_time",
    "elapsed_sec",     # endTime - startTime（纯推理耗时）
    "wait_sec",        # startTime - enqueueTime（排队耗时）
    "answer",
    "think",
    "khObj",
    "skillsResult",
    "error",
    "poll_http_status",
    "poll_biz_status",
    "poll_biz_message",
    "polled_at",
]


def load_questions(path: str) -> list[str]:
    """读取测试集"问题"列，CSV / xlsx 都支持。"""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        try:
            import openpyxl  # type: ignore
        except ImportError as e:
            raise RuntimeError("读取 xlsx 需要 openpyxl，请先 pip install openpyxl") from e
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        header = [
            str(c.value).strip() if c.value is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        q_col = "问题" if "问题" in header else ("question" if "question" in header else None)
        if q_col is None:
            raise ValueError(f"输入 xlsx 缺少'问题'或'question'列，实际表头={header}")
        q_idx = header.index(q_col)
        items: list[str] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if q_idx >= len(row):
                continue
            v = row[q_idx]
            if v is None:
                continue
            q = str(v).strip()
            if q:
                items.append(q)
        return items

    items: list[str] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fn = reader.fieldnames or []
        q_key = "问题" if "问题" in fn else ("question" if "question" in fn else None)
        if q_key is None:
            raise ValueError(f"输入 CSV 缺少'问题'或'question'列，实际表头={reader.fieldnames}")
        for row in reader:
            q = (row.get(q_key) or "").strip()
            if q:
                items.append(q)
    return items


# ---------------------------------------------------------------- IO 工具

def _blank_row(idx: int, question: str) -> dict:
    return {k: "" for k in OUTPUT_FIELDS} | {"idx": idx, "question": question}


def _rewrite_csv(path: str, rows: list[dict]) -> None:
    """全量重写 CSV（用于 poll 阶段更新结果）。"""
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in OUTPUT_FIELDS})
        f.flush()
        try:
            os.fsync(f.fileno())
        except OSError:
            pass
    os.replace(tmp, path)


def _append_row(writer: csv.DictWriter, fout, lock: threading.Lock, row: dict) -> None:
    with lock:
        writer.writerow({k: row.get(k, "") for k in OUTPUT_FIELDS})
        fout.flush()
        try:
            os.fsync(fout.fileno())
        except OSError:
            pass


# ---------------------------------------------------------------- submit

def submit_task(
    base_url: str,
    policy_id: str,
    question: str,
    timeout: float,
    extra_payload: dict | None = None,
) -> dict:
    """调用 /api/reason/submit，返回 {taskId, submit_http_status, submit_biz_status, submit_error}。"""
    url = base_url.rstrip("/") + "/submit"
    payload = {"question": question, "policyId": policy_id}
    if extra_payload:
        for k, v in extra_payload.items():
            if k not in payload:
                payload[k] = v
    out = {
        "taskId": "",
        "submit_http_status": "",
        "submit_biz_status": "",
        "submit_error": "",
    }
    try:
        resp = requests.post(
            url,
            json=payload,
            timeout=timeout,
            headers={"Content-Type": "application/json"},
        )
        out["submit_http_status"] = resp.status_code
        try:
            body = resp.json()
        except Exception:
            out["submit_error"] = f"响应非 JSON: {resp.text[:500]}"
            return out
        out["submit_biz_status"] = body.get("status_code")
        data = body.get("data") or {}
        if isinstance(data, dict):
            out["taskId"] = data.get("taskId", "") or ""
        if body.get("status_code") != 200:
            out["submit_error"] = body.get("message", "") or f"submit 非 200"
    except requests.Timeout:
        out["submit_error"] = f"submit 超时(>{timeout}s)"
    except Exception as e:
        out["submit_error"] = f"{type(e).__name__}: {e}"
    return out


def run_submit_phase(
    questions: list[str],
    output_csv: str,
    base_url: str,
    policy_id: str,
    concurrency: int,
    timeout: float,
    extra_payload: dict | None,
) -> list[dict]:
    """并发 submit 并实时把 taskId 落盘到 CSV（追加写）。返回所有行（保持 idx 升序）。"""
    total = len(questions)
    os.makedirs(os.path.dirname(os.path.abspath(output_csv)) or ".", exist_ok=True)
    lock = threading.Lock()
    counter = {"done": 0, "ok": 0, "fail": 0}
    rows_by_idx: dict[int, dict] = {}

    started = time.time()
    with open(output_csv, "w", encoding="utf-8-sig", newline="") as fout:
        writer = csv.DictWriter(fout, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        fout.flush()

        def worker(idx: int, question: str) -> dict:
            res = submit_task(base_url, policy_id, question, timeout, extra_payload)
            row = _blank_row(idx, question)
            row.update(res)
            row["submitted_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            row["status"] = "pending" if res["taskId"] else "submit_failed"
            _append_row(writer, fout, lock, row)
            rows_by_idx[idx] = row
            with lock:
                counter["done"] += 1
                if res["taskId"]:
                    counter["ok"] += 1
                else:
                    counter["fail"] += 1
                print(
                    f"[submit {counter['done']}/{total}] idx={idx} "
                    f"taskId={res['taskId'] or '-'} "
                    f"biz={res['submit_biz_status']} "
                    f"err={res['submit_error'][:80]}",
                    flush=True,
                )
            return row

        with ThreadPoolExecutor(max_workers=concurrency) as pool:
            futures = [pool.submit(worker, i + 1, q) for i, q in enumerate(questions)]
            for fut in as_completed(futures):
                try:
                    fut.result()
                except Exception as e:
                    print(f"[WARN] submit worker 异常: {type(e).__name__}: {e}", flush=True)

    elapsed = time.time() - started
    print(
        f"\n[SUBMIT DONE] 总数={total} 成功={counter['ok']} 失败={counter['fail']} "
        f"耗时={elapsed:.1f}s"
    )

    rows = [rows_by_idx[i + 1] for i in range(total) if (i + 1) in rows_by_idx]
    rows.sort(key=lambda r: int(r["idx"]))
    return rows


# ---------------------------------------------------------------- poll

def poll_one(
    base_url: str,
    task_id: str,
    poll_interval: float,
    poll_timeout: float,
    req_timeout: float,
) -> dict:
    """轮询一个 taskId 直到 done/failed 或到 poll_timeout。

    返回 dict，字段与 OUTPUT_FIELDS 中 poll 相关字段对齐。
    """
    url = base_url.rstrip("/") + f"/result/{task_id}"
    out: dict = {
        "status": "",
        "enqueue_time": "",
        "start_time": "",
        "end_time": "",
        "elapsed_sec": "",
        "wait_sec": "",
        "answer": "",
        "think": "",
        "khObj": "",
        "skillsResult": "",
        "error": "",
        "poll_http_status": "",
        "poll_biz_status": "",
        "poll_biz_message": "",
        "polled_at": "",
    }
    deadline = time.time() + poll_timeout
    last_status = ""

    while True:
        try:
            resp = requests.get(url, timeout=req_timeout)
            out["poll_http_status"] = resp.status_code
            try:
                body = resp.json()
            except Exception:
                last_status = f"non_json_http_{resp.status_code}"
                out["error"] = f"response not JSON: {resp.text[:300]}"
                if time.time() >= deadline:
                    out["status"] = "failed"
                    break
                time.sleep(poll_interval)
                continue
            out["poll_biz_status"] = body.get("status_code")
            out["poll_biz_message"] = body.get("message", "")
            if body.get("status_code") == 404:
                out["status"] = "not_found"
                out["error"] = body.get("message", "") or "taskId 不存在/已过期"
                break
            if body.get("status_code") != 200:
                # 503 等临时错误：继续重试
                last_status = f"biz={body.get('status_code')}"
            else:
                out["error"] = ""
                data = body.get("data") or {}
                status = data.get("status", "") or ""
                out["status"] = status
                out["enqueue_time"] = data.get("enqueueTime", "") or ""
                out["start_time"] = data.get("startTime", "") or ""
                out["end_time"] = data.get("endTime", "") or ""
                if isinstance(out["end_time"], (int, float)) and isinstance(
                    out["start_time"], (int, float)
                ):
                    out["elapsed_sec"] = round(out["end_time"] - out["start_time"], 3)
                if isinstance(out["start_time"], (int, float)) and isinstance(
                    out["enqueue_time"], (int, float)
                ):
                    out["wait_sec"] = round(out["start_time"] - out["enqueue_time"], 3)
                if status == "done":
                    result = data.get("result") or {}
                    if isinstance(result, dict):
                        out["answer"] = result.get("answer", "") or ""
                        out["think"] = result.get("think", "") or ""
                        out["khObj"] = result.get("khObj", "") or ""
                        skills = result.get("skillsResult") or {}
                        if skills:
                            try:
                                out["skillsResult"] = json.dumps(skills, ensure_ascii=False)
                            except Exception:
                                out["skillsResult"] = str(skills)
                    break
                if status == "failed":
                    out["error"] = data.get("error", "") or "failed"
                    break
                last_status = status
        except requests.Timeout:
            last_status = f"poll_timeout(>{req_timeout}s)"
        except Exception as e:
            last_status = f"{type(e).__name__}: {e}"

        if time.time() >= deadline:
            out["status"] = out["status"] or "timeout"
            out["error"] = out["error"] or f"等待超时(>{poll_timeout}s) last={last_status}"
            break
        time.sleep(poll_interval)

    out["polled_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return out


def run_poll_phase(
    rows: list[dict],
    output_csv: str,
    base_url: str,
    concurrency: int,
    poll_interval: float,
    poll_timeout: float,
    req_timeout: float,
) -> None:
    """并发对已有 taskId 的行进行轮询；完成一条即回写整个 CSV。"""
    todo = [r for r in rows if (r.get("taskId") or "").strip() and r.get("status") not in ("done", "failed", "not_found")]
    total = len(todo)
    if total == 0:
        print("[POLL] 无待查询任务")
        _rewrite_csv(output_csv, rows)
        return

    print(f"[POLL] 待查询任务数={total}, 并发={concurrency}, 间隔={poll_interval}s, 超时={poll_timeout}s")
    lock = threading.Lock()
    counter = {"done": 0, "ok": 0, "fail": 0}
    overall_start = time.time()

    def worker(row: dict) -> None:
        task_id = row["taskId"]
        res = poll_one(base_url, task_id, poll_interval, poll_timeout, req_timeout)
        with lock:
            row.update(res)
            counter["done"] += 1
            if res["status"] == "done" and not res["error"]:
                counter["ok"] += 1
            else:
                counter["fail"] += 1
            _rewrite_csv(output_csv, rows)
            preview = (res["answer"] or res["error"] or "").replace("\n", " ")[:60]
            print(
                f"[poll {counter['done']}/{total}] idx={row['idx']} "
                f"status={res['status']} "
                f"reason={res['elapsed_sec']}s wait={res['wait_sec']}s | {preview}",
                flush=True,
            )

    with ThreadPoolExecutor(max_workers=concurrency) as pool:
        futures = [pool.submit(worker, r) for r in todo]
        for fut in as_completed(futures):
            try:
                fut.result()
            except Exception as e:
                print(f"[WARN] poll worker 异常: {type(e).__name__}: {e}", flush=True)

    elapsed = time.time() - overall_start
    print(
        f"\n[POLL DONE] 总数={total} 成功={counter['ok']} 失败={counter['fail']} "
        f"耗时={elapsed:.1f}s 平均={elapsed/total:.2f}s/题"
    )


def load_existing_csv(path: str) -> list[dict]:
    if not os.path.isfile(path):
        return []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = [dict(r) for r in reader]
    for r in rows:
        for k in OUTPUT_FIELDS:
            r.setdefault(k, "")
    rows.sort(key=lambda r: int(r.get("idx") or 0))
    return rows


def _clear_submit_and_poll_fields(row: dict) -> None:
    """保留 idx、question，清空 taskId 及之后的提交/轮询结果，用于重跑失败样本。"""
    for k in OUTPUT_FIELDS:
        if k in ("idx", "question"):
            continue
        row[k] = ""


def run_retry_submit_failed_phase(
    rows: list[dict],
    output_csv: str,
    base_url: str,
    policy_id: str,
    concurrency: int,
    submit_timeout: float,
    extra_payload: dict | None,
) -> None:
    failed = [r for r in rows if (r.get("status") or "").strip() != "done"]
    if not failed:
        print("[RETRY] 无非 done 行，无需重新提交")
        return
    total = len(failed)
    print(
        f"[RETRY] 重新提交失败行={total}, idx={[int(r['idx']) for r in failed]}, "
        f"并发={concurrency}"
    )
    lock = threading.Lock()
    counter = {"done": 0, "ok": 0, "fail": 0}

    def worker(r: dict) -> None:
        _clear_submit_and_poll_fields(r)
        q = (r.get("question") or "").strip()
        if not q:
            r["status"] = "submit_failed"
            r["submit_error"] = "question 为空"
            with lock:
                counter["done"] += 1
                counter["fail"] += 1
                _rewrite_csv(output_csv, rows)
            return
        res = submit_task(base_url, policy_id, q, submit_timeout, extra_payload)
        r.update(res)
        r["submitted_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        r["status"] = "pending" if res["taskId"] else "submit_failed"
        with lock:
            rows.sort(key=lambda x: int(x.get("idx") or 0))
            _rewrite_csv(output_csv, rows)
            counter["done"] += 1
            if res["taskId"]:
                counter["ok"] += 1
            else:
                counter["fail"] += 1
            print(
                f"[retry-submit {counter['done']}/{total}] idx={r['idx']} "
                f"taskId={res['taskId'] or '-'} biz={res['submit_biz_status']} "
                f"err={(res['submit_error'] or '')[:80]}",
                flush=True,
            )

    with ThreadPoolExecutor(max_workers=concurrency) as pool:
        futures = [pool.submit(worker, r) for r in failed]
        for fut in as_completed(futures):
            try:
                fut.result()
            except Exception as e:
                print(f"[WARN] retry submit 异常: {type(e).__name__}: {e}", flush=True)

    print(f"[RETRY SUBMIT DONE] 成功 taskId={counter['ok']} 失败={counter['fail']}")


# ---------------------------------------------------------------- main

def main() -> int:
    parser = argparse.ArgumentParser(description="async submit + poll 跑 ncp 测试集")
    parser.add_argument("--input", default=DEFAULT_INPUT, help="输入 CSV/xlsx 路径")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="结果 CSV 路径")
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL, help="/api/reason 基础地址（不带 /submit）")
    parser.add_argument("--policy-id", default=DEFAULT_POLICY_ID, help="policyId")

    parser.add_argument("--submit-concurrency", type=int, default=10, help="submit 并发")
    parser.add_argument("--submit-timeout", type=float, default=30.0, help="submit 单次 HTTP 超时(s)")

    parser.add_argument("--poll-concurrency", type=int, default=20, help="poll 并发（受服务端 worker 数天花板约束）")
    parser.add_argument("--poll-interval", type=float, default=2.0, help="轮询间隔(s)")
    parser.add_argument("--poll-timeout", type=float, default=3600.0, help="单任务等待上限(s)")
    parser.add_argument("--poll-req-timeout", type=float, default=30.0, help="result 查询单次 HTTP 超时(s)")

    parser.add_argument("--limit", type=int, default=0, help="只跑前 N 条（0=全部）")
    parser.add_argument("--submit-only", action="store_true", help="只提交，不查结果")
    parser.add_argument("--poll-only", action="store_true", help="不提交，仅从 --output CSV 中读取 taskId 继续查询")
    parser.add_argument(
        "--retry-failed",
        action="store_true",
        help="从 --output 读取结果表，对 status≠done 的行重新 submit（并默认继续 poll）；与 --poll-only 互斥",
    )

    parser.add_argument("--vendor", default="", help="LLM vendor；为空走服务端默认")
    parser.add_argument("--model", default="", help="LLM model；为空走服务端默认")
    parser.add_argument(
        "--answer-system-prompt-file",
        default="",
        help="answerSystemPrompt 文件路径（utf-8）",
    )
    parser.add_argument(
        "--extra-json",
        default="",
        help='额外透传字段，JSON 对象，例如 \'{"reduceMaxPartDepth":4}\'',
    )

    args = parser.parse_args()

    if args.submit_only and args.poll_only:
        print("[ERROR] --submit-only 与 --poll-only 不能同时指定")
        return 1
    if args.retry_failed and args.poll_only:
        print("[ERROR] --retry-failed 与 --poll-only 不能同时指定")
        return 1

    extra_payload: dict = {}
    if args.answer_system_prompt_file:
        with open(args.answer_system_prompt_file, "r", encoding="utf-8") as fp:
            extra_payload["answerSystemPrompt"] = fp.read()
    if args.vendor:
        extra_payload["vendor"] = args.vendor
    if args.model:
        extra_payload["model"] = args.model
    if args.extra_json:
        try:
            extras = json.loads(args.extra_json)
        except Exception as e:
            print(f"[ERROR] --extra-json 解析失败: {e}")
            return 1
        if not isinstance(extras, dict):
            print("[ERROR] --extra-json 必须是 JSON 对象")
            return 1
        extra_payload.update(extras)

    print(f"[INFO] base-url : {args.base_url}")
    print(f"[INFO] policyId : {args.policy_id}")
    print(f"[INFO] output   : {args.output}")
    if extra_payload:
        keys = ", ".join(f"{k}({len(str(v))}字符)" for k, v in extra_payload.items())
        print(f"[INFO] 额外入参 : {keys}")

    rows: list[dict]

    if args.retry_failed:
        rows = load_existing_csv(args.output)
        if not rows:
            print(f"[ERROR] --retry-failed 模式下找不到或空文件: {args.output}")
            return 1
        print(f"[INFO] --retry-failed：已读 {len(rows)} 行，准备重提交失败行")
        run_retry_submit_failed_phase(
            rows,
            args.output,
            args.base_url,
            args.policy_id,
            args.submit_concurrency,
            args.submit_timeout,
            extra_payload or None,
        )
        _rewrite_csv(args.output, rows)
    elif args.poll_only:
        rows = load_existing_csv(args.output)
        if not rows:
            print(f"[ERROR] --poll-only 模式下找不到或空文件: {args.output}")
            return 1
        print(f"[INFO] 从现有 CSV 读取 {len(rows)} 行，进入 poll 阶段")
    else:
        questions = load_questions(args.input)
        if args.limit > 0:
            questions = questions[: args.limit]
        if not questions:
            print("[ERROR] 输入文件中没有有效问题")
            return 1
        print(f"[INFO] 输入文件 : {args.input}")
        print(f"[INFO] 问题总数 : {len(questions)}")
        rows = run_submit_phase(
            questions,
            args.output,
            args.base_url,
            args.policy_id,
            args.submit_concurrency,
            args.submit_timeout,
            extra_payload or None,
        )
        # submit 阶段已按"完成即追加"写入 CSV，这里再统一按 idx 升序重写一次，便于阅读
        _rewrite_csv(args.output, rows)

    if args.submit_only:
        print(f"[DONE] submit-only 模式结束，taskId 已落盘：{args.output}")
        print("       后续可用 --poll-only --output <同一文件> 继续查结果")
        return 0

    run_poll_phase(
        rows,
        args.output,
        args.base_url,
        args.poll_concurrency,
        args.poll_interval,
        args.poll_timeout,
        args.poll_req_timeout,
    )

    ok = sum(1 for r in rows if r.get("status") == "done" and not r.get("error"))
    fail = sum(1 for r in rows if r.get("status") != "done" or r.get("error"))
    print(f"\n[ALL DONE] 成功={ok} 失败={fail} 总={len(rows)}")
    print(f"[ALL DONE] 结果文件: {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
